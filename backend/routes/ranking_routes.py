"""
Ranking Routes
Handles AHP and other ranking algorithms
"""

from flask import Blueprint, request, jsonify
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from utils.ahp import ahp_ranking, calculate_consistency_ratio, create_pairwise_matrix
from utils.file_manager import load_excel, save_excel
import pandas as pd
from openpyxl.styles import Alignment

ranking_bp = Blueprint('ranking', __name__)
PROJECT_FILE = 'project.xlsx'  # Ranking data is now stored in project.xlsx
OLD_RANKING_FILE = 'ranking.xlsx'  # Old file name for migration

def migrate_ranking_data():
    """Migrate ranking data from ranking.xlsx to project.xlsx if needed"""
    from pathlib import Path
    DATA_DIR = Path(__file__).parent.parent.parent / 'data'
    old_filepath = DATA_DIR / OLD_RANKING_FILE
    new_filepath = DATA_DIR / PROJECT_FILE
    
    # Only migrate if old file exists and hasn't been migrated yet
    if not old_filepath.exists():
        return False
    
    try:
        # Load ranking data from old file
        xls = pd.ExcelFile(old_filepath)
        ranking_columns = []
        criteria_weights = {}
        alternatives_scores_df = None
        ranking_result = None
        
        # Load ranking columns
        if 'RankingColumns' in xls.sheet_names:
            df_columns = pd.read_excel(old_filepath, sheet_name='RankingColumns')
            ranking_columns = [
                str(row['column']).strip() 
                for _, row in df_columns.iterrows() 
                if 'column' in row and pd.notna(row['column'])
            ]
        
        # Load criteria weights
        if 'CriteriaWeights' in xls.sheet_names:
            df_weights = pd.read_excel(old_filepath, sheet_name='CriteriaWeights')
            for _, row in df_weights.iterrows():
                if 'criteria' in row and 'weight' in row:
                    try:
                        criteria = str(row['criteria']).strip()
                        weight = row['weight']
                        if pd.notna(criteria) and pd.notna(weight):
                            criteria_weights[criteria] = float(weight)
                    except (ValueError, TypeError):
                        continue
        
        # Load alternatives scores
        if 'AlternativesScores' in xls.sheet_names:
            alternatives_scores_df = pd.read_excel(old_filepath, sheet_name='AlternativesScores')
        
        # Load ranking result
        if 'RankingResult' in xls.sheet_names:
            ranking_result = pd.read_excel(old_filepath, sheet_name='RankingResult')
        
        # Load existing project data
        project_rows = []
        if new_filepath.exists():
            try:
                project_df = pd.read_excel(new_filepath, sheet_name='Project')
                project_rows = project_df.to_dict('records')
            except:
                try:
                    project_df = pd.read_excel(new_filepath, sheet_name='Sheet1')
                    project_rows = project_df.to_dict('records')
                except:
                    project_rows = []
        
        # Save all data to project.xlsx
        with pd.ExcelWriter(new_filepath, engine='openpyxl') as writer:
            # Save project data
            if project_rows:
                pd.DataFrame(project_rows).to_excel(writer, sheet_name='Project', index=False)
            else:
                pd.DataFrame().to_excel(writer, sheet_name='Project', index=False)
            
            # Save ranking data
            if ranking_columns:
                df_columns = pd.DataFrame([{'column': col} for col in ranking_columns])
                df_columns.to_excel(writer, sheet_name='RankingColumns', index=False)
            
            if criteria_weights:
                df_weights = pd.DataFrame([
                    {'criteria': k, 'weight': v} 
                    for k, v in criteria_weights.items()
                ])
                df_weights.to_excel(writer, sheet_name='CriteriaWeights', index=False)
            
            if alternatives_scores_df is not None and not alternatives_scores_df.empty:
                alternatives_scores_df.to_excel(writer, sheet_name='AlternativesScores', index=False)
            
            if ranking_result is not None and not ranking_result.empty:
                ranking_result.to_excel(writer, sheet_name='RankingResult', index=False)
            
            # Apply center alignment to all cells in all sheets
            workbook = writer.book
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Delete old file after successful migration
        try:
            old_filepath.unlink()
        except Exception:
            pass  # If deletion fails, continue anyway
        
        return True
    except Exception as e:
        print(f"Error migrating ranking data: {str(e)}")
        return False

@ranking_bp.route('/ahp', methods=['POST'])
def perform_ahp():
    """Perform AHP ranking"""
    data = request.json
    criteria_weights = data.get('criteriaWeights', {})
    alternatives_scores = data.get('alternativesScores', {})
    
    if not criteria_weights or not alternatives_scores:
        return jsonify({'success': False, 'error': 'Criteria weights and alternatives scores required'}), 400
    
    try:
        result = ahp_ranking(criteria_weights, alternatives_scores)
        return jsonify({
            'success': True,
            'data': result
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@ranking_bp.route('/pairwise-matrix', methods=['POST'])
def create_pairwise():
    """Create pairwise comparison matrix"""
    data = request.json
    scores = data.get('scores', [])
    
    try:
        matrix = create_pairwise_matrix(scores)
        cr, ci, max_eigenvalue = calculate_consistency_ratio(matrix)
        
        return jsonify({
            'success': True,
            'matrix': matrix,
            'consistency_ratio': float(cr),
            'consistency_index': float(ci),
            'max_eigenvalue': float(max_eigenvalue),
            'is_consistent': cr < 0.1
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@ranking_bp.route('/other', methods=['POST'])
def other_algorithms():
    """Placeholder for other ranking algorithms"""
    return jsonify({
        'success': False,
        'message': 'TOPSIS, VIKOR, and other algorithms need further development'
    })

@ranking_bp.route('/get', methods=['GET'])
def get_ranking():
    """Get current ranking data from project.xlsx"""
    try:
        # Migrate data from old ranking.xlsx if it exists
        migrate_ranking_data()
        
        from pathlib import Path
        DATA_DIR = Path(__file__).parent.parent.parent / 'data'
        filepath = DATA_DIR / PROJECT_FILE
        
        # Initialize default structure
        ranking_data = {
            'criteriaWeights': {},
            'alternativesScores': {},
            'rankingResult': None,
            'columns': [],
            'groups': []
        }
        
        if filepath.exists():
            try:
                # Load from multiple sheets
                xls = pd.ExcelFile(filepath)
                criteria_weights = {}
                alternatives_scores = {}
                ranking_result = None
                ranking_columns = []
                ranking_groups = []
                
                # Sheet 0: Ranking Columns (optional)
                if 'RankingColumns' in xls.sheet_names:
                    df_columns = pd.read_excel(filepath, sheet_name='RankingColumns')
                    for _, row in df_columns.iterrows():
                        if 'column' in row:
                            col = str(row['column']).strip()
                            if pd.notna(col) and col:
                                ranking_columns.append(col)
                
                # Sheet 1: Ranking Groups (optional)
                if 'RankingGroups' in xls.sheet_names:
                    try:
                        df_groups = pd.read_excel(filepath, sheet_name='RankingGroups')
                        for _, row in df_groups.iterrows():
                            if 'group_name' in row and 'columns' in row:
                                try:
                                    group_name = str(row['group_name']).strip()
                                    columns_str = str(row['columns']).strip()
                                    if pd.notna(group_name) and pd.notna(columns_str):
                                        # Parse columns string (e.g., "col1,col2,col3")
                                        columns = [col.strip() for col in columns_str.split(',') if col.strip()]
                                        ranking_groups.append({
                                            'name': group_name,
                                            'columns': columns
                                        })
                                except (ValueError, TypeError) as e:
                                    print(f"Error parsing group row: {e}")
                                    continue
                    except Exception as e:
                        print(f"Error loading RankingGroups sheet: {e}")
                        # Continue without loading groups
                
                # Sheet 2: Criteria Weights
                if 'CriteriaWeights' in xls.sheet_names:
                    df_weights = pd.read_excel(filepath, sheet_name='CriteriaWeights')
                    for _, row in df_weights.iterrows():
                        if 'criteria' in row and 'weight' in row:
                            try:
                                criteria = str(row['criteria']).strip()
                                weight = row['weight']
                                if pd.notna(criteria) and pd.notna(weight):
                                    criteria_weights[criteria] = float(weight)
                            except (ValueError, TypeError):
                                continue
                
                # Sheet 3: Alternatives Scores
                if 'AlternativesScores' in xls.sheet_names:
                    df_scores = pd.read_excel(filepath, sheet_name='AlternativesScores')
                    # Check if it's new format (table with alternative column + criteria columns)
                    # or old format (alternative, criteria, score columns)
                    if (
                        'alternative' in df_scores.columns
                        and 'criteria' in df_scores.columns
                        and 'score' in df_scores.columns
                    ):
                        # Old format: 3 columns (alternative, criteria, score)
                        for _, row in df_scores.iterrows():
                            if 'alternative' in row and 'criteria' in row and 'score' in row:
                                try:
                                    alt = str(row['alternative']).strip()
                                    criteria = str(row['criteria']).strip()
                                    score = row['score']
                                    if pd.notna(alt) and pd.notna(criteria) and pd.notna(score):
                                        score_val = float(score)
                                        if alt not in alternatives_scores:
                                            alternatives_scores[alt] = {}
                                        alternatives_scores[alt][criteria] = score_val
                                except (ValueError, TypeError):
                                    continue
                    else:
                        # New format: alternative column + criteria columns
                        if 'alternative' in df_scores.columns:
                            # Get all criteria columns (all columns except 'alternative')
                            criteria_cols = [col for col in df_scores.columns if col != 'alternative']
                            for _, row in df_scores.iterrows():
                                try:
                                    alt = str(row['alternative']).strip()
                                    if pd.notna(alt) and alt:
                                        alternatives_scores[alt] = {}
                                        for criteria in criteria_cols:
                                            score = row[criteria]
                                            if pd.notna(score):
                                                try:
                                                    score_val = float(score)
                                                    alternatives_scores[alt][criteria] = score_val
                                                except (ValueError, TypeError):
                                                    continue
                                except (ValueError, TypeError):
                                    continue
                
                # Sheet 4: Ranking Result (optional)
                if 'RankingResult' in xls.sheet_names:
                    df_result = pd.read_excel(filepath, sheet_name='RankingResult')
                    ranking_result = {
                        'ranking': [
                            {'alternative': str(row['alternative']), 'score': float(row['score'])}
                            for _, row in df_result.iterrows()
                            if 'alternative' in row and 'score' in row and pd.notna(row['alternative']) and pd.notna(row['score'])
                        ]
                    }
                    if not ranking_result['ranking']:
                        ranking_result = None
                
                ranking_data = {
                    'criteriaWeights': criteria_weights,
                    'alternativesScores': alternatives_scores,
                    'rankingResult': ranking_result,
                    'columns': ranking_columns,
                    'groups': ranking_groups
                }
            except Exception as e:
                # If parsing fails, return empty structure
                print(f"Error parsing ranking file: {str(e)}")
                ranking_data = {
                    'criteriaWeights': {},
                    'alternativesScores': {},
                    'rankingResult': None,
                    'columns': [],
                    'groups': []
                }
        else:
            # File doesn't exist - return empty structure
            print(f"Ranking file does not exist: {filepath}")
            ranking_data = {
                'criteriaWeights': {},
                'alternativesScores': {},
                'rankingResult': None,
                'columns': [],
                'groups': []
            }
        
        return jsonify({
            'success': True,
            'data': ranking_data
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'data': {
                'criteriaWeights': {},
                'alternativesScores': {},
                'rankingResult': None,
                'columns': [],
                'groups': []
            }
        }), 500

@ranking_bp.route('/delete', methods=['POST'])
def delete_ranking():
    """Delete ranking data from project.xlsx (clear ranking sheets, preserve project data)"""
    try:
        from pathlib import Path
        DATA_DIR = Path(__file__).parent.parent.parent / 'data'
        filepath = DATA_DIR / PROJECT_FILE
        
        # Load project data if file exists
        project_rows = []
        if filepath.exists():
            try:
                project_df = pd.read_excel(filepath, sheet_name='Project')
                project_rows = project_df.to_dict('records')
            except:
                try:
                    project_df = pd.read_excel(filepath, sheet_name='Sheet1')
                    project_rows = project_df.to_dict('records')
                except:
                    project_rows = []
        
        # Save project data with empty ranking sheets
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Preserve project data
            if project_rows:
                pd.DataFrame(project_rows).to_excel(writer, sheet_name='Project', index=False)
            else:
                pd.DataFrame().to_excel(writer, sheet_name='Project', index=False)
            
            # Clear ranking sheets
            pd.DataFrame(columns=['column']).to_excel(writer, sheet_name='RankingColumns', index=False)
            pd.DataFrame(columns=['group_name', 'columns']).to_excel(writer, sheet_name='RankingGroups', index=False)
            pd.DataFrame(columns=['criteria', 'weight']).to_excel(writer, sheet_name='CriteriaWeights', index=False)
            pd.DataFrame(columns=['alternative']).to_excel(writer, sheet_name='AlternativesScores', index=False)
            pd.DataFrame(columns=['alternative', 'score']).to_excel(writer, sheet_name='RankingResult', index=False)
            
            # Apply center alignment to all cells in all sheets
            workbook = writer.book
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        return jsonify({
            'success': True,
            'message': 'Ranking data cleared successfully'
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@ranking_bp.route('/cell/update', methods=['POST'])
def update_ranking_cell():
    """Update a single cell in ranking (score or weight).

    This implementation works directly on the AlternativesScores sheet as a table
    (row = alternative, column = criteria) to avoid shuffling or losing existing
    scores when a single cell changes.
    """
    try:
        data = request.json
        cell_type = data.get('type')  # 'score' or 'weight'
        alternative = data.get('alternative', '')
        criteria = data.get('criteria', '')
        value = data.get('value', '')

        DATA_DIR = Path(__file__).parent.parent.parent / 'data'
        DATA_DIR.mkdir(exist_ok=True)
        filepath = DATA_DIR / PROJECT_FILE

        # --- Load existing meta data (columns, weights, result) ---
        ranking_columns = []
        criteria_weights = {}
        ranking_result = None
        df_scores = None

        if filepath.exists():
            try:
                xls = pd.ExcelFile(filepath)

                # Ranking columns
                if 'RankingColumns' in xls.sheet_names:
                    df_columns = pd.read_excel(filepath, sheet_name='RankingColumns')
                    for _, row in df_columns.iterrows():
                        if 'column' in row and pd.notna(row['column']):
                            col_name = str(row['column']).strip()
                            if col_name:
                                ranking_columns.append(col_name)

                # Criteria weights
                if 'CriteriaWeights' in xls.sheet_names:
                    df_weights = pd.read_excel(filepath, sheet_name='CriteriaWeights')
                    for _, row in df_weights.iterrows():
                        if 'criteria' in row and 'weight' in row:
                            try:
                                crit_name = str(row['criteria']).strip()
                                weight = row['weight']
                                if pd.notna(crit_name) and pd.notna(weight):
                                    criteria_weights[crit_name] = float(weight)
                            except (ValueError, TypeError):
                                continue

                # Ranking result (preserve if exists)
                if 'RankingResult' in xls.sheet_names:
                    df_result = pd.read_excel(filepath, sheet_name='RankingResult')
                    tmp_ranking = [
                        {
                            'alternative': str(row['alternative']),
                            'score': float(row['score'])
                        }
                        for _, row in df_result.iterrows()
                        if 'alternative' in row and 'score' in row
                        and pd.notna(row['alternative']) and pd.notna(row['score'])
                    ]
                    if tmp_ranking:
                        ranking_result = {'ranking': tmp_ranking}

                # Alternatives scores as table
                if 'AlternativesScores' in xls.sheet_names:
                    df_scores_raw = pd.read_excel(filepath, sheet_name='AlternativesScores')
                    if (
                        'alternative' in df_scores_raw.columns and
                        'criteria' in df_scores_raw.columns and
                        'score' in df_scores_raw.columns
                    ):
                        # Old format: pivot to wide table
                        df_scores = (
                            df_scores_raw
                            .dropna(subset=['alternative'])
                            .pivot_table(
                                index='alternative',
                                columns='criteria',
                                values='score',
                                aggfunc='first'
                            )
                            .reset_index()
                        )
                        df_scores.columns.name = None
                    else:
                        # New format already (wide table)
                        df_scores = df_scores_raw
            except Exception as e:
                print(f"Error loading ranking file: {str(e)}")

        # Ensure we have a DataFrame for AlternativesScores
        if df_scores is None:
            base_cols = ['alternative'] + list(ranking_columns)
            df_scores = pd.DataFrame(columns=base_cols)
        else:
            # Ensure 'alternative' exists and is first
            if 'alternative' not in df_scores.columns:
                df_scores.insert(0, 'alternative', None)
            else:
                other_cols = [c for c in df_scores.columns if c != 'alternative']
                df_scores = df_scores[['alternative'] + other_cols]

        # Make sure all known ranking columns exist in df_scores
        for col in ranking_columns:
            if col not in df_scores.columns:
                df_scores[col] = pd.NA

        # --- Initialize missing alternatives based on project rows (if available) ---
        # This ensures that when a new ranking file is created and the user starts
        # entering scores cell-by-cell, we still keep a full row per project row
        # instead of only creating the specific alternative that was edited.
        project_rows = []
        if filepath.exists():
            try:
                project_df = pd.read_excel(filepath, sheet_name='Project')
                project_rows = project_df.to_dict('records')
            except:
                try:
                    project_df = pd.read_excel(filepath, sheet_name='Sheet1')
                    project_rows = project_df.to_dict('records')
                except:
                    project_rows = []

        if project_rows:
            # Build a set of existing alternative keys in the current scores table
            existing_alternatives = set(
                df_scores['alternative']
                .dropna()
                .astype(str)
                .str.strip()
                .tolist()
            )

            # For each project row, ensure there is a corresponding alternative row
            for idx, row in enumerate(project_rows):
                row_no = row.get('rowNo') or idx + 1
                try:
                    row_no_int = int(row_no)
                except (TypeError, ValueError):
                    row_no_int = idx + 1
                alt_key = f"Alternative {row_no_int}"

                if alt_key not in existing_alternatives:
                    new_row = {col_name: pd.NA for col_name in df_scores.columns}
                    new_row['alternative'] = alt_key
                    df_scores = pd.concat(
                        [df_scores, pd.DataFrame([new_row])],
                        ignore_index=True
                    )
                    existing_alternatives.add(alt_key)

        # --- Apply the specific cell update ---
        if cell_type == 'weight':
            if criteria:
                try:
                    weight_value = float(value) if value != '' else 0
                    criteria_weights[criteria] = weight_value
                except (ValueError, TypeError):
                    criteria_weights[criteria] = 0

                # Ensure criteria appears as a column / ranking column
                if criteria not in ranking_columns:
                    ranking_columns.append(criteria)
                if criteria not in df_scores.columns:
                    df_scores[criteria] = pd.NA

        elif cell_type == 'score':
            if alternative and criteria:
                # Ensure criteria column exists
                if criteria not in df_scores.columns:
                    df_scores[criteria] = pd.NA
                    if criteria not in ranking_columns:
                        ranking_columns.append(criteria)

                alt_str = str(alternative).strip()
                mask = df_scores['alternative'].astype(str).str.strip() == alt_str

                if not mask.any():
                    # Add a new row for this alternative
                    new_row = {col_name: pd.NA for col_name in df_scores.columns}
                    new_row['alternative'] = alt_str
                    df_scores = pd.concat(
                        [df_scores, pd.DataFrame([new_row])],
                        ignore_index=True
                    )
                    mask = df_scores['alternative'].astype(str).str.strip() == alt_str

                try:
                    score_value = float(value) if value != '' else 0
                except (ValueError, TypeError):
                    score_value = 0

                df_scores.loc[mask, criteria] = score_value

        # Normalize column order: alternative + sorted criteria columns
        criteria_cols_in_df = [c for c in df_scores.columns if c != 'alternative']
        all_criteria_cols = sorted(set(criteria_cols_in_df + list(ranking_columns)))
        if df_scores.shape[0] > 0:
            df_scores = df_scores[['alternative'] + all_criteria_cols]
        else:
            df_scores = pd.DataFrame(columns=['alternative'] + all_criteria_cols)

        # --- Save back to Excel - preserve project data and sync AlternativesScores ---
        # Load existing project data
        project_rows = []
        if filepath.exists():
            try:
                project_df = pd.read_excel(filepath, sheet_name='Project')
                project_rows = project_df.to_dict('records')
            except:
                try:
                    project_df = pd.read_excel(filepath, sheet_name='Sheet1')
                    project_rows = project_df.to_dict('records')
                except:
                    project_rows = []
        
        # Sync AlternativesScores with Project rows - ensure same number of rows
        if project_rows and df_scores is not None:
            num_project_rows = len(project_rows)
            
            # Get existing alternatives as dict for quick lookup
            existing_alternatives_dict = {}
            if not df_scores.empty:
                for _, alt_row in df_scores.iterrows():
                    alt_key = str(alt_row.get('alternative', '')).strip()
                    if alt_key:
                        existing_alternatives_dict[alt_key] = alt_row.to_dict()
            
            # Build new alternatives list matching project rows
            new_alternatives = []
            for idx, row in enumerate(project_rows):
                row_no = row.get('rowNo', idx + 1)
                alt_key = f"Alternative {row_no}"
                
                # Use existing data if available, otherwise create new
                if alt_key in existing_alternatives_dict:
                    alt_data = existing_alternatives_dict[alt_key].copy()
                    alt_data['alternative'] = alt_key  # Ensure alternative key is correct
                    new_alternatives.append(alt_data)
                else:
                    # Create new alternative row
                    new_alt = {'alternative': alt_key}
                    for col in all_criteria_cols:
                        new_alt[col] = pd.NA
                    new_alternatives.append(new_alt)
            
            # Create DataFrame from synced alternatives
            if new_alternatives:
                df_scores = pd.DataFrame(new_alternatives)
                # Ensure all columns exist
                required_cols = ['alternative'] + all_criteria_cols
                for col in required_cols:
                    if col not in df_scores.columns:
                        df_scores[col] = pd.NA
                df_scores = df_scores[required_cols]
            else:
                df_scores = pd.DataFrame(columns=['alternative'] + all_criteria_cols)
        elif project_rows and df_scores is None:
            # No alternatives scores yet, create based on project rows
            new_alternatives = []
            for idx, row in enumerate(project_rows):
                row_no = row.get('rowNo', idx + 1)
                alt_key = f"Alternative {row_no}"
                new_alt = {'alternative': alt_key}
                for col in all_criteria_cols:
                    new_alt[col] = pd.NA
                new_alternatives.append(new_alt)
            df_scores = pd.DataFrame(new_alternatives)
            if all_criteria_cols:
                df_scores = df_scores[['alternative'] + all_criteria_cols]
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Preserve project data with column order
            if project_rows:
                # Get column order from existing file if available
                if filepath.exists():
                    try:
                        df_existing = pd.read_excel(filepath, sheet_name='Project')
                        existing_col_order = list(df_existing.columns)
                        # Reorder project_rows to match existing column order
                        df_project = pd.DataFrame(project_rows)
                        # Ensure all columns from existing file are present
                        for col in existing_col_order:
                            if col not in df_project.columns:
                                df_project[col] = pd.NA
                        # Reorder to match existing order, then add any new columns
                        # BUT ensure rowNo is always first
                        ordered_cols = [c for c in existing_col_order if c in df_project.columns]
                        new_cols = [c for c in df_project.columns if c not in ordered_cols]
                        # Make sure rowNo is first
                        final_cols = []
                        if 'rowNo' in df_project.columns:
                            final_cols.append('rowNo')
                        for col in ordered_cols + new_cols:
                            if col != 'rowNo':
                                final_cols.append(col)
                        df_project = df_project[final_cols]
                        df_project.to_excel(writer, sheet_name='Project', index=False)
                    except:
                        # Fallback: ensure rowNo is first
                        df_project = pd.DataFrame(project_rows)
                        if 'rowNo' in df_project.columns:
                            cols = ['rowNo'] + [c for c in df_project.columns if c != 'rowNo']
                            df_project = df_project[cols]
                        df_project.to_excel(writer, sheet_name='Project', index=False)
                else:
                    # No existing file, ensure rowNo is first
                    df_project = pd.DataFrame(project_rows)
                    if 'rowNo' in df_project.columns:
                        cols = ['rowNo'] + [c for c in df_project.columns if c != 'rowNo']
                        df_project = df_project[cols]
                    df_project.to_excel(writer, sheet_name='Project', index=False)
            else:
                pd.DataFrame().to_excel(writer, sheet_name='Project', index=False)
            
            # Save ranking data
            if ranking_columns:
                df_columns = pd.DataFrame([{'column': col} for col in ranking_columns])
                df_columns.to_excel(writer, sheet_name='RankingColumns', index=False)

            if criteria_weights:
                df_weights = pd.DataFrame(
                    [{'criteria': k, 'weight': v} for k, v in criteria_weights.items()]
                )
                df_weights.to_excel(writer, sheet_name='CriteriaWeights', index=False)

            # AlternativesScores sheet (always wide format, synced with Project)
            if df_scores is not None:
                if df_scores.empty:
                    pd.DataFrame(columns=['alternative'] + all_criteria_cols).to_excel(
                        writer, sheet_name='AlternativesScores', index=False
                    )
                else:
                    df_scores.to_excel(writer, sheet_name='AlternativesScores', index=False)

            if ranking_result and ranking_result.get('ranking'):
                df_result = pd.DataFrame(ranking_result['ranking'])
                df_result.to_excel(writer, sheet_name='RankingResult', index=False)
            
            # Apply center alignment to all cells in all sheets
            workbook = writer.book
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

        return jsonify({'success': True, 'message': 'Cell updated successfully'})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@ranking_bp.route('/update', methods=['POST'])
def update_ranking():
    """Update ranking data and save to Excel"""
    try:
        data = request.json
        criteria_weights = data.get('criteriaWeights', {})
        alternatives_scores = data.get('alternativesScores', {})
        ranking_result = data.get('rankingResult', None)
        ranking_columns = data.get('columns', [])
        ranking_groups = data.get('groups', [])
        
        # Debug logging
        print(f"Received data:")
        print(f"  - columns: {ranking_columns}")
        print(f"  - groups: {ranking_groups}")
        print(f"  - criteria_weights: {criteria_weights}")
        
        DATA_DIR = Path(__file__).parent.parent.parent / 'data'
        DATA_DIR.mkdir(exist_ok=True)
        filepath = DATA_DIR / PROJECT_FILE
        
        # Load existing project data to preserve it
        project_rows = []
        if filepath.exists():
            try:
                project_df = pd.read_excel(filepath, sheet_name='Project')
                project_rows = project_df.to_dict('records')
            except:
                try:
                    project_df = pd.read_excel(filepath, sheet_name='Sheet1')
                    project_rows = project_df.to_dict('records')
                except:
                    project_rows = []
        
        # Save as Excel with multiple sheets - preserve project data and column order (rowNo must be first)
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Preserve project data with column order
            if project_rows:
                # Get column order from existing file if available
                if filepath.exists():
                    try:
                        df_existing = pd.read_excel(filepath, sheet_name='Project')
                        existing_col_order = list(df_existing.columns)
                        # Reorder project_rows to match existing column order
                        df_project = pd.DataFrame(project_rows)
                        # Ensure all columns from existing file are present
                        for col in existing_col_order:
                            if col not in df_project.columns:
                                df_project[col] = pd.NA
                        # Reorder to match existing order, then add any new columns
                        # BUT ensure rowNo is always first
                        ordered_cols = [c for c in existing_col_order if c in df_project.columns]
                        new_cols = [c for c in df_project.columns if c not in ordered_cols]
                        # Make sure rowNo is first
                        final_cols = []
                        if 'rowNo' in df_project.columns:
                            final_cols.append('rowNo')
                        for col in ordered_cols + new_cols:
                            if col != 'rowNo':
                                final_cols.append(col)
                        df_project = df_project[final_cols]
                        df_project.to_excel(writer, sheet_name='Project', index=False)
                    except:
                        # Fallback: ensure rowNo is first
                        df_project = pd.DataFrame(project_rows)
                        if 'rowNo' in df_project.columns:
                            cols = ['rowNo'] + [c for c in df_project.columns if c != 'rowNo']
                            df_project = df_project[cols]
                        df_project.to_excel(writer, sheet_name='Project', index=False)
                else:
                    # No existing file, ensure rowNo is first
                    df_project = pd.DataFrame(project_rows)
                    if 'rowNo' in df_project.columns:
                        cols = ['rowNo'] + [c for c in df_project.columns if c != 'rowNo']
                        df_project = df_project[cols]
                    df_project.to_excel(writer, sheet_name='Project', index=False)
            else:
                pd.DataFrame().to_excel(writer, sheet_name='Project', index=False)
            
            # Sheet 0: Ranking Columns
            if ranking_columns:
                df_columns = pd.DataFrame([
                    {'column': col}
                    for col in ranking_columns
                ])
                df_columns.to_excel(writer, sheet_name='RankingColumns', index=False)
            
            # Sheet 1: Ranking Groups
            print(f"Saving groups: {ranking_groups}")
            if ranking_groups:
                # Convert groups to a flat structure for Excel
                groups_data = []
                for group in ranking_groups:
                    if 'name' in group and 'columns' in group and group['columns']:
                        groups_data.append({
                            'group_name': group['name'],
                            'columns': ','.join(group['columns'])
                        })
                print(f"Groups data to save: {groups_data}")
                if groups_data:
                    df_groups = pd.DataFrame(groups_data)
                    print(f"Saving DataFrame: {df_groups}")
                    df_groups.to_excel(writer, sheet_name='RankingGroups', index=False)
            else:
                # Create empty RankingGroups sheet if no groups
                print("Creating empty RankingGroups sheet")
                pd.DataFrame(columns=['group_name', 'columns']).to_excel(writer, sheet_name='RankingGroups', index=False)
            
            # Sheet 2: Criteria Weights
            if criteria_weights:
                df_weights = pd.DataFrame([
                    {'criteria': k, 'weight': v}
                    for k, v in criteria_weights.items()
                ])
                df_weights.to_excel(writer, sheet_name='CriteriaWeights', index=False)
            
            # Sheet 3: Alternatives Scores - New format: table with alternative column + criteria columns
            # Sync with Project rows to ensure same number of rows
            if alternatives_scores:
                # Get all unique criteria from all alternatives
                all_criteria = set()
                for alt, scores in alternatives_scores.items():
                    if scores and isinstance(scores, dict):
                        all_criteria.update(scores.keys())
                
                # Also include criteria from ranking_columns if available
                if ranking_columns:
                    all_criteria.update(ranking_columns)
                
                all_criteria = sorted(list(all_criteria))
                
                # Build table structure: alternative column + criteria columns
                # First, create a dict from alternatives_scores for quick lookup
                alternatives_dict = {}
                for alt, scores in alternatives_scores.items():
                    alternatives_dict[str(alt)] = scores if scores and isinstance(scores, dict) else {}
                
                # Sync with project rows
                scores_rows = []
                if project_rows:
                    # Use project rows as the source of truth for row count and order
                    for idx, proj_row in enumerate(project_rows):
                        row_no = proj_row.get('rowNo', idx + 1)
                        alt_key = f"Alternative {row_no}"
                        
                        # Get scores for this alternative
                        scores = alternatives_dict.get(alt_key, {})
                        row = {'alternative': alt_key}
                        for criteria in all_criteria:
                            score_value = scores.get(criteria, '')
                            if score_value is not None and score_value != '':
                                try:
                                    row[criteria] = float(score_value)
                                except (ValueError, TypeError):
                                    row[criteria] = pd.NA
                            else:
                                row[criteria] = pd.NA
                        scores_rows.append(row)
                else:
                    # No project rows, use alternatives_scores as is
                    for alt in sorted(alternatives_scores.keys()):
                        row = {'alternative': str(alt)}
                        scores = alternatives_scores.get(alt, {})
                        for criteria in all_criteria:
                            score_value = scores.get(criteria, '')
                            if score_value is not None and score_value != '':
                                try:
                                    row[criteria] = float(score_value)
                                except (ValueError, TypeError):
                                    row[criteria] = pd.NA
                            else:
                                row[criteria] = pd.NA
                        scores_rows.append(row)
                
                if scores_rows:
                    df_scores = pd.DataFrame(scores_rows)
                    df_scores.to_excel(writer, sheet_name='AlternativesScores', index=False)
                else:
                    # Empty structure with alternative column
                    df_scores = pd.DataFrame(columns=['alternative'] + list(all_criteria))
                    df_scores.to_excel(writer, sheet_name='AlternativesScores', index=False)
            elif project_rows and ranking_columns:
                # No alternatives_scores but we have project rows and ranking columns
                # Create empty alternatives scores matching project rows
                scores_rows = []
                for idx, proj_row in enumerate(project_rows):
                    row_no = proj_row.get('rowNo', idx + 1)
                    alt_key = f"Alternative {row_no}"
                    row = {'alternative': alt_key}
                    for criteria in ranking_columns:
                        row[criteria] = pd.NA
                    scores_rows.append(row)
                df_scores = pd.DataFrame(scores_rows)
                df_scores.to_excel(writer, sheet_name='AlternativesScores', index=False)
            
            # Sheet 4: Ranking Result
            if ranking_result and ranking_result.get('ranking'):
                df_result = pd.DataFrame(ranking_result['ranking'])
                df_result.to_excel(writer, sheet_name='RankingResult', index=False)
            
            # Apply center alignment to all cells in all sheets
            workbook = writer.book
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        return jsonify({
            'success': True,
            'message': 'Ranking updated successfully',
            'data': {
                'criteriaWeights': criteria_weights,
                'alternativesScores': alternatives_scores,
                'rankingResult': ranking_result,
                'columns': ranking_columns,
                'groups': ranking_groups
            }
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

