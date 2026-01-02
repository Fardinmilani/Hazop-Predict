"""
Report Routes
Handles report generation and statistics
"""

from flask import Blueprint, request, jsonify, send_file
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from utils.report_generator import (
    generate_statistics, generate_correlation_heatmap,
    generate_bar_chart, generate_scatter_plot
)
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch
from io import BytesIO
import base64
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

report_bp = Blueprint('report', __name__)

@report_bp.route('/statistics', methods=['POST'])
def get_statistics():
    """Get statistical summary"""
    data = request.json
    project_data = data.get('data', [])
    include_plots = data.get('includeDistributionPlots', False)
    
    try:
        stats = generate_statistics(project_data, include_distribution_plots=include_plots)
        return jsonify({
            'success': True,
            'data': stats
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@report_bp.route('/visualizations', methods=['POST'])
def get_visualizations():
    """Get visualizations"""
    data = request.json
    project_data = data.get('data', [])
    viz_type = data.get('type', 'heatmap')  # heatmap, bar, scatter, distribution
    include_categorical = data.get('includeCategorical', False)  # For correlation heatmap
    
    try:
        if viz_type == 'heatmap':
            image = generate_correlation_heatmap(project_data, include_categorical=include_categorical)
        elif viz_type == 'bar':
            column = data.get('column')
            image = generate_bar_chart(project_data, column)
        elif viz_type == 'scatter':
            x_col = data.get('xColumn')
            y_col = data.get('yColumn')
            image = generate_scatter_plot(project_data, x_col, y_col)
        elif viz_type == 'distribution':
            column = data.get('column')
            if not column:
                return jsonify({'success': False, 'error': 'Column name required'}), 400
            # Get column data
            df = pd.DataFrame(project_data)
            if 'rowNo' in df.columns:
                df = df.drop(columns=['rowNo'])
            if column not in df.columns:
                return jsonify({'success': False, 'error': 'Column not found'}), 400
            col_data = df[column].dropna()
            if len(col_data) == 0:
                return jsonify({'success': False, 'error': 'No data available'}), 400
            # Fit distribution
            from utils.report_generator import fit_distribution, generate_distribution_plot
            distribution_fit = fit_distribution(col_data.values)
            if distribution_fit:
                image = generate_distribution_plot(col_data.values, column, distribution_fit)
            else:
                return jsonify({'success': False, 'error': 'Could not fit distribution'}), 400
        else:
            return jsonify({'success': False, 'error': 'Invalid visualization type'}), 400
        
        if image is None:
            return jsonify({'success': False, 'error': 'Could not generate visualization'}), 400
        
        return jsonify({
            'success': True,
            'image': image
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@report_bp.route('/export-excel', methods=['POST'])
def export_excel():
    """Export report as Excel - returns file in memory"""
    data = request.json
    project_data = data.get('data', [])
    filename = data.get('filename', 'report.xlsx')
    
    try:
        # Generate statistics with distribution plots
        stats = generate_statistics(project_data, include_distribution_plots=True)
        
        # Create Excel in memory
        output = BytesIO()
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Sheet 1: Data
        ws_data = wb.create_sheet("Data")
        df_data = pd.DataFrame(project_data)
        if 'rowNo' in df_data.columns:
            df_data = df_data.drop(columns=['rowNo'])
        
        # Write headers
        headers = list(df_data.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data
        for row_idx, row_data in enumerate(df_data.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                ws_data.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            max_length = max(len(str(header)), 
                           max([len(str(row[col_idx-1])) for row in df_data.values] + [10]))
            ws_data.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
        
        # Sheet 2: Numeric Statistics
        if stats['numeric_stats']:
            ws_stats = wb.create_sheet("Numeric Statistics")
            stats_headers = ['Column', 'Mean', 'Std', 'Min', 'Max', 'Median', 'Q25', 'Q75', 'Distribution', 'P-Value', 'Parameters']
            for col_idx, header in enumerate(stats_headers, 1):
                cell = ws_stats.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for row_idx, (col, stat) in enumerate(stats['numeric_stats'].items(), 2):
                ws_stats.cell(row=row_idx, column=1, value=col)
                ws_stats.cell(row=row_idx, column=2, value=round(stat['mean'], 2))
                ws_stats.cell(row=row_idx, column=3, value=round(stat['std'], 2))
                ws_stats.cell(row=row_idx, column=4, value=round(stat['min'], 2))
                ws_stats.cell(row=row_idx, column=5, value=round(stat['max'], 2))
                ws_stats.cell(row=row_idx, column=6, value=round(stat['median'], 2))
                ws_stats.cell(row=row_idx, column=7, value=round(stat['q25'], 2))
                ws_stats.cell(row=row_idx, column=8, value=round(stat['q75'], 2))
                
                if stat.get('distribution'):
                    dist_info = stat['distribution']
                    ws_stats.cell(row=row_idx, column=9, value=dist_info['distribution'])
                    ws_stats.cell(row=row_idx, column=10, value=round(dist_info.get('pvalue', 0), 4))
                    # Format parameters as string
                    params_str = ', '.join([f"{p:.4f}" for p in dist_info['params']])
                    ws_stats.cell(row=row_idx, column=11, value=params_str)
                else:
                    ws_stats.cell(row=row_idx, column=9, value='N/A')
                    ws_stats.cell(row=row_idx, column=10, value='N/A')
                    ws_stats.cell(row=row_idx, column=11, value='N/A')
            
            # Auto-adjust column widths
            for col_idx in range(1, len(stats_headers) + 1):
                ws_stats.column_dimensions[get_column_letter(col_idx)].width = 15
        
        # Sheet 3: Categorical Statistics
        if stats.get('categorical_stats'):
            ws_cat = wb.create_sheet("Categorical Statistics")
            cat_headers = ['Column', 'Unique Count', 'Most Frequent Values']
            for col_idx, header in enumerate(cat_headers, 1):
                cell = ws_cat.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for row_idx, (col, stat) in enumerate(stats['categorical_stats'].items(), 2):
                ws_cat.cell(row=row_idx, column=1, value=col)
                ws_cat.cell(row=row_idx, column=2, value=stat['unique_count'])
                freq_str = ', '.join([f"{val}: {count}" for val, count in stat['most_frequent'][:5]])
                ws_cat.cell(row=row_idx, column=3, value=freq_str)
            
            # Auto-adjust column widths
            for col_idx in range(1, len(cat_headers) + 1):
                ws_cat.column_dimensions[get_column_letter(col_idx)].width = 30
        
        # Sheet 4: Summary
        ws_summary = wb.create_sheet("Summary", 0)  # Insert at beginning
        ws_summary.cell(row=1, column=1, value="HAZOP Analysis Report")
        ws_summary.cell(row=1, column=1).font = Font(size=16, bold=True)
        ws_summary.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws_summary.cell(row=4, column=1, value=f"Total Records: {stats['count']}")
        ws_summary.cell(row=5, column=1, value=f"Total Columns: {len(stats['columns'])}")
        ws_summary.cell(row=6, column=1, value=f"Numeric Columns: {len(stats['numeric_stats'])}")
        ws_summary.cell(row=7, column=1, value=f"Categorical Columns: {len(stats.get('categorical_stats', {}))}")
        ws_summary.column_dimensions['A'].width = 40
        
        wb.save(output)
        output.seek(0)
        
        # Save to data directory temporarily for download
        DATA_DIR = Path(__file__).parent.parent.parent / 'data'
        filepath = DATA_DIR / filename
        filepath.parent.mkdir(exist_ok=True)
        with open(filepath, 'wb') as f:
            f.write(output.getvalue())
        
        return jsonify({
            'success': True,
            'filepath': str(filepath),
            'filename': filename
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@report_bp.route('/download-excel', methods=['GET'])
def download_excel():
    """Download Excel file"""
    filename = request.args.get('filename', 'report.xlsx')
    DATA_DIR = Path(__file__).parent.parent.parent / 'data'
    filepath = DATA_DIR / filename
    
    if not filepath.exists():
        return jsonify({'success': False, 'error': 'File not found'}), 404
    
    return send_file(str(filepath), as_attachment=True, download_name=filename)

@report_bp.route('/export-pdf', methods=['POST'])
def export_pdf():
    """Export report as PDF"""
    data = request.json
    project_data = data.get('data', [])
    filename = data.get('filename', 'report.pdf')
    
    try:
        stats = generate_statistics(project_data, include_distribution_plots=True)
        
        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        styles = getSampleStyleSheet()
        
        # Import PIL for image handling
        from reportlab.platypus import Image as RLImage
        from PIL import Image as PILImage
        
        # Title
        title = Paragraph("HAZOP Analysis Report", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 0.2*inch))
        
        # Summary
        story.append(Paragraph("Report Summary", styles['Heading1']))
        summary_data = [
            ['Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Total Records', str(stats['count'])],
            ['Total Columns', str(len(stats['columns']))],
            ['Numeric Columns', str(len(stats['numeric_stats']))],
            ['Categorical Columns', str(len(stats.get('categorical_stats', {})))]
        ]
        summary_table = Table(summary_data, colWidths=[2*inch, 4*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Numeric statistics table
        if stats['numeric_stats']:
            story.append(Paragraph("Numeric Statistics", styles['Heading2']))
            data_table = [['Column', 'Mean', 'Std', 'Min', 'Max', 'Median', 'Q25', 'Q75', 'Distribution', 'P-Value']]
            for col, stat in stats['numeric_stats'].items():
                if stat.get('distribution'):
                    dist_info = stat['distribution']
                    dist_name = dist_info['distribution']
                    p_value = f"{dist_info.get('pvalue', 0):.4f}"
                else:
                    dist_name = 'N/A'
                    p_value = 'N/A'
                
                data_table.append([
                    col[:12] + '...' if len(col) > 12 else col,
                    f"{stat['mean']:.2f}",
                    f"{stat['std']:.2f}",
                    f"{stat['min']:.2f}",
                    f"{stat['max']:.2f}",
                    f"{stat['median']:.2f}",
                    f"{stat['q25']:.2f}",
                    f"{stat['q75']:.2f}",
                    dist_name[:10],
                    p_value
                ])
            
            table = Table(data_table, colWidths=[1*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.9*inch, 0.8*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            story.append(table)
            story.append(Spacer(1, 0.2*inch))
            
            # Add distribution parameters table
            story.append(Paragraph("Distribution Parameters", styles['Heading3']))
            params_table_data = [['Column', 'Distribution', 'Parameters', 'P-Value']]
            for col, stat in stats['numeric_stats'].items():
                if stat.get('distribution'):
                    dist_info = stat['distribution']
                    params_str = ', '.join([f"{p:.4f}" for p in dist_info['params']])
                    params_table_data.append([
                        col[:15] + '...' if len(col) > 15 else col,
                        dist_info['distribution'],
                        params_str[:40] + '...' if len(params_str) > 40 else params_str,
                        f"{dist_info.get('pvalue', 0):.4f}"
                    ])
            
            if len(params_table_data) > 1:
                params_table = Table(params_table_data, colWidths=[1.5*inch, 1.2*inch, 3*inch, 0.8*inch])
                params_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                story.append(params_table)
            
            story.append(PageBreak())
            
            # Note about distribution plots
            story.append(Paragraph("Distribution Fit Plots", styles['Heading2']))
            story.append(Paragraph(
                "Note: Distribution plots are available in the web interface. "
                "Click on 'View Plot' button next to each column in the Statistics Summary to view the distribution fit.",
                styles['Normal']
            ))
        
        # Categorical statistics
        if stats.get('categorical_stats'):
            story.append(Paragraph("Categorical Statistics", styles['Heading2']))
            cat_table_data = [['Column', 'Unique Count', 'Most Frequent Values']]
            for col, stat in stats['categorical_stats'].items():
                freq_str = ', '.join([f"{val}: {count}" for val, count in stat['most_frequent'][:5]])
                cat_table_data.append([
                    col[:20] + '...' if len(col) > 20 else col,
                    str(stat['unique_count']),
                    freq_str[:50] + '...' if len(freq_str) > 50 else freq_str
                ])
            
            cat_table = Table(cat_table_data, colWidths=[2*inch, 1*inch, 3*inch])
            cat_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            story.append(cat_table)
        
        doc.build(story)
        buffer.seek(0)
        
        # Save PDF temporarily for download
        DATA_DIR = Path(__file__).parent.parent.parent / 'data'
        filepath = DATA_DIR / filename
        filepath.parent.mkdir(exist_ok=True)
        with open(filepath, 'wb') as f:
            f.write(buffer.getvalue())
        
        return jsonify({
            'success': True,
            'filepath': str(filepath),
            'filename': filename
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@report_bp.route('/download-pdf', methods=['GET'])
def download_pdf():
    """Download PDF file"""
    filename = request.args.get('filename', 'report.pdf')
    DATA_DIR = Path(__file__).parent.parent.parent / 'data'
    filepath = DATA_DIR / filename
    
    if not filepath.exists():
        return jsonify({'success': False, 'error': 'File not found'}), 404
    
    return send_file(str(filepath), as_attachment=True, download_name=filename)

