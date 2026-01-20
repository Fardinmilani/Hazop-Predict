"""
File Routes
Handles file operations: New, Open, Save, Save As
"""

from flask import Blueprint, request, jsonify, send_file
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from utils.file_manager import (
    save_json, load_json, save_csv, load_csv,
    save_excel, load_excel, list_files
)
import os
import pandas as pd
import io
from openpyxl.styles import Alignment

file_bp = Blueprint('file', __name__)

# Data directory path
DATA_DIR = Path(__file__).parent.parent.parent / 'data'

@file_bp.route('/new', methods=['POST'])
def new_file():
    """Create a new empty project"""
    return jsonify({
        'success': True,
        'message': 'New project created',
        'data': {
            'rows': [],
            'columns': []
        }
    })

@file_bp.route('/open', methods=['POST'])
def open_file():
    """Open an existing project file - receives file content or filename"""
    import io
    
    try:
        # Check if file is uploaded or filename is provided
        if 'file' in request.files:
            # File uploaded from file picker
            file = request.files['file']
            if file.filename == '':
                return jsonify({'success': False, 'error': 'No file selected'}), 400
            
            filename = file.filename
            file_content = file.read()
            file_stream = io.BytesIO(file_content)
        elif request.is_json and request.json and 'filename' in request.json:
            # Filename provided (from saved files list)
            filename = request.json['filename']
            
            # Handle special files (project.xlsx and library.json)
            if filename == 'project.xlsx':
                from pathlib import Path
                DATA_DIR = Path(__file__).parent.parent.parent / 'data'
                filepath = DATA_DIR / 'project.xlsx'
                
                if not filepath.exists():
                    project = {'rows': [], 'columns': []}
                else:
                    try:
                        # Load project data from 'Project' or 'Sheet1' sheet
                        project_rows = []
                        project_columns = []
                        try:
                            project_df = pd.read_excel(filepath, sheet_name='Project')
                            project_rows = project_df.to_dict('records')
                            if len(project_rows) > 0:
                                project_columns = [col for col in list(project_rows[0].keys()) if col != 'rowNo']
                        except:
                            try:
                                project_df = pd.read_excel(filepath, sheet_name='Sheet1')
                                project_rows = project_df.to_dict('records')
                                if len(project_rows) > 0:
                                    project_columns = [col for col in list(project_rows[0].keys()) if col != 'rowNo']
                            except:
                                project_rows = []
                        
                        # Load ranking data if available
                        xls = pd.ExcelFile(filepath)
                        criteria_weights = {}
                        alternatives_scores = {}
                        ranking_result = None
                        ranking_columns = []
                        ranking_groups = []
                        severity_values = {}
                        recommendations = {}
                        optimum_weights = {}
                        criteria_directions = {}
                        
                        # Load RankingColumns
                        if 'RankingColumns' in xls.sheet_names:
                            df_columns = pd.read_excel(filepath, sheet_name='RankingColumns')
                            for _, row in df_columns.iterrows():
                                if 'column' in row:
                                    col = str(row['column']).strip()
                                    if pd.notna(col) and col:
                                        ranking_columns.append(col)
                        
                        # Load RankingGroups
                        if 'RankingGroups' in xls.sheet_names:
                            try:
                                df_groups = pd.read_excel(filepath, sheet_name='RankingGroups')
                                for _, row in df_groups.iterrows():
                                    if 'group_name' in row and 'columns' in row:
                                        try:
                                            group_name = str(row['group_name']).strip()
                                            columns_str = str(row['columns']).strip()
                                            if pd.notna(group_name) and pd.notna(columns_str):
                                                columns = [col.strip() for col in columns_str.split(',') if col.strip()]
                                                ranking_groups.append({
                                                    'name': group_name,
                                                    'columns': columns
                                                })
                                        except (ValueError, TypeError):
                                            continue
                            except Exception:
                                pass
                        
                        # Load CriteriaWeights with directions
                        if 'CriteriaWeights' in xls.sheet_names:
                            df_weights = pd.read_excel(filepath, sheet_name='CriteriaWeights')
                            for _, row in df_weights.iterrows():
                                if 'criteria' in row and 'weight' in row:
                                    criteria = str(row['criteria']).strip()
                                    weight = row['weight']
                                    if pd.notna(criteria) and pd.notna(weight):
                                        criteria_weights[criteria] = float(weight)
                                    # Load direction if available
                                    if 'direction' in row and pd.notna(row['direction']):
                                        criteria_directions[criteria] = str(row['direction']).strip()
                        
                        # Load AlternativesScores
                        if 'AlternativesScores' in xls.sheet_names:
                            df_scores = pd.read_excel(filepath, sheet_name='AlternativesScores')
                            if 'alternative' in df_scores.columns:
                                criteria_cols = [col for col in df_scores.columns if col != 'alternative']
                                for _, row in df_scores.iterrows():
                                    alt = str(row['alternative']).strip()
                                    if pd.notna(alt) and alt:
                                        alternatives_scores[alt] = {}
                                        for criteria in criteria_cols:
                                            score = row[criteria]
                                            if pd.notna(score):
                                                try:
                                                    alternatives_scores[alt][criteria] = float(score)
                                                except (ValueError, TypeError):
                                                    continue
                        
                        # Load RankingResult
                        if 'RankingResult' in xls.sheet_names:
                            df_result = pd.read_excel(filepath, sheet_name='RankingResult')
                            ranking_result = {
                                'ranking': [
                                    {'alternative': row['alternative'], 'score': float(row['score'])}
                                    for _, row in df_result.iterrows()
                                    if 'alternative' in row and 'score' in row and pd.notna(row['alternative']) and pd.notna(row['score'])
                                ]
                            }
                            if not ranking_result['ranking']:
                                ranking_result = None
                        
                        # Load RiskAssessment
                        if 'RiskAssessment' in xls.sheet_names:
                            df_risk = pd.read_excel(filepath, sheet_name='RiskAssessment')
                            for _, row in df_risk.iterrows():
                                if 'alternative' in row and pd.notna(row['alternative']):
                                    alt = str(row['alternative']).strip()
                                    if 'severity' in row and pd.notna(row['severity']):
                                        try:
                                            severity_values[alt] = float(row['severity'])
                                        except (ValueError, TypeError):
                                            pass
                                    if 'recommendation' in row and pd.notna(row['recommendation']):
                                        recommendations[alt] = str(row['recommendation']).strip()
                                    if 'optimum_weight' in row and pd.notna(row['optimum_weight']):
                                        try:
                                            optimum_weights[alt] = float(row['optimum_weight'])
                                        except (ValueError, TypeError):
                                            pass
                        
                        # Combine project and ranking data
                        project = {
                            'rows': project_rows,
                            'columns': project_columns
                        }
                        
                        # Add ranking data if available
                        if criteria_weights or alternatives_scores or ranking_result or ranking_columns:
                            project.update({
                                'criteriaWeights': criteria_weights,
                                'alternativesScores': alternatives_scores,
                                'rankingResult': ranking_result,
                                'rankingColumns': ranking_columns,
                                'groups': ranking_groups,
                                'severityValues': severity_values,
                                'recommendations': recommendations,
                                'optimumWeights': optimum_weights,
                                'criteriaDirections': criteria_directions
                            })
                    except Exception as e:
                        project = {'rows': [], 'columns': []}
                
                return jsonify({
                    'success': True,
                    'data': project,
                    'filename': filename
                })
            elif filename == 'library.json':
                from utils.file_manager import load_json
                library = load_json('library.json')
                if library is None:
                    library = {'headers': []}
                return jsonify({
                    'success': True,
                    'data': library,
                    'filename': filename
                })
            
            # Check if file exists in data directory
            filepath = DATA_DIR / filename
            if not filepath.exists():
                return jsonify({'success': False, 'error': 'File not found'}), 404
            
            file_content = filepath.read_bytes()
            file_stream = io.BytesIO(file_content)
        else:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        
        file_extension = filename.split('.')[-1].lower()
        file_name_lower = filename.lower()
        
        # Read file content based on extension
        if file_extension == 'xlsx':
            # Check if it's a ranking file or project file with ranking data
            try:
                file_stream.seek(0)
                xls = pd.ExcelFile(file_stream)
                
                # Initialize project data
                project_rows = []
                project_columns = []
                
                # Try to load project data from 'Project' sheet or first sheet
                if 'Project' in xls.sheet_names:
                    file_stream.seek(0)
                    try:
                        project_df = pd.read_excel(file_stream, sheet_name='Project')
                        project_rows = project_df.to_dict('records')
                        if len(project_rows) > 0:
                            project_columns = [col for col in list(project_rows[0].keys()) if col != 'rowNo']
                    except:
                        pass
                elif len(xls.sheet_names) > 0:
                    # Try first sheet if 'Project' sheet doesn't exist
                    # But only if it's not a ranking sheet
                    first_sheet = xls.sheet_names[0]
                    if first_sheet not in ['RankingColumns', 'CriteriaWeights', 'AlternativesScores', 'RankingResult']:
                        file_stream.seek(0)
                        try:
                            project_df = pd.read_excel(file_stream, sheet_name=first_sheet)
                            project_rows = project_df.to_dict('records')
                            if len(project_rows) > 0:
                                project_columns = [col for col in list(project_rows[0].keys()) if col != 'rowNo']
                        except:
                            pass
                
                # Check if it has ranking sheets
                has_ranking_sheets = 'CriteriaWeights' in xls.sheet_names or 'AlternativesScores' in xls.sheet_names
                
                if has_ranking_sheets or 'project.xlsx' in file_name_lower:
                    # It has ranking data - load ranking sheets
                    criteria_weights = {}
                    alternatives_scores = {}
                    ranking_result = None
                    ranking_columns = []
                    
                    # Sheet 0: Ranking Columns (optional)
                    if 'RankingColumns' in xls.sheet_names:
                        file_stream.seek(0)
                        df_columns = pd.read_excel(file_stream, sheet_name='RankingColumns')
                        for _, row in df_columns.iterrows():
                            if 'column' in row:
                                col = str(row['column']).strip()
                                if pd.notna(col) and col:
                                    ranking_columns.append(col)
                    
                    # Sheet 0.5: Ranking Groups (optional)
                    ranking_groups = []
                    if 'RankingGroups' in xls.sheet_names:
                        file_stream.seek(0)
                        try:
                            df_groups = pd.read_excel(file_stream, sheet_name='RankingGroups')
                            for _, row in df_groups.iterrows():
                                if 'group_name' in row and 'columns' in row:
                                    try:
                                        group_name = str(row['group_name']).strip()
                                        columns_str = str(row['columns']).strip()
                                        if pd.notna(group_name) and pd.notna(columns_str):
                                            # Parse columns string (e.g., "col1,col2,col3")
                                            columns = [col.strip() for col in columns_str.split(',') if col.strip()]
                                            ranking_groups.append({
                                                'name': group_name,
                                                'columns': columns
                                            })
                                    except (ValueError, TypeError):
                                        continue
                        except Exception:
                            pass
                    
                    # Sheet 1: Criteria Weights (with directions)
                    criteria_directions = {}
                    if 'CriteriaWeights' in xls.sheet_names:
                        file_stream.seek(0)
                        df_weights = pd.read_excel(file_stream, sheet_name='CriteriaWeights')
                        for _, row in df_weights.iterrows():
                            if 'criteria' in row and 'weight' in row:
                                criteria = str(row['criteria']).strip()
                                weight = row['weight']
                                if pd.notna(criteria) and pd.notna(weight):
                                    criteria_weights[criteria] = float(weight)
                                # Load direction if available
                                if 'direction' in row and pd.notna(row['direction']):
                                    criteria_directions[criteria] = str(row['direction']).strip()
                    
                    # Sheet 2: Alternatives Scores
                    if 'AlternativesScores' in xls.sheet_names:
                        file_stream.seek(0)
                        df_scores = pd.read_excel(file_stream, sheet_name='AlternativesScores')
                        # Check if it's new format (table with alternative column + criteria columns)
                        # or old format (alternative, criteria, score columns)
                        if 'alternative' in df_scores.columns and 'criteria' in df_scores.columns and 'score' in df_scores.columns:
                            # Old format: 3 columns (alternative, criteria, score)
                            for _, row in df_scores.iterrows():
                                if 'alternative' in row and 'criteria' in row and 'score' in row:
                                    alt = str(row['alternative']).strip()
                                    criteria = str(row['criteria']).strip()
                                    score = row['score']
                                    if pd.notna(alt) and pd.notna(criteria) and pd.notna(score):
                                        if alt not in alternatives_scores:
                                            alternatives_scores[alt] = {}
                                        alternatives_scores[alt][criteria] = float(score)
                        else:
                            # New format: alternative column + criteria columns
                            if 'alternative' in df_scores.columns:
                                # Get all criteria columns (all columns except 'alternative')
                                criteria_cols = [col for col in df_scores.columns if col != 'alternative']
                                for _, row in df_scores.iterrows():
                                    alt = str(row['alternative']).strip()
                                    if pd.notna(alt) and alt:
                                        alternatives_scores[alt] = {}
                                        for criteria in criteria_cols:
                                            score = row[criteria]
                                            if pd.notna(score):
                                                try:
                                                    alternatives_scores[alt][criteria] = float(score)
                                                except (ValueError, TypeError):
                                                    continue
                    
                    # Sheet 3: Ranking Result (optional)
                    if 'RankingResult' in xls.sheet_names:
                        file_stream.seek(0)
                        df_result = pd.read_excel(file_stream, sheet_name='RankingResult')
                        ranking_result = {
                            'ranking': [
                                {'alternative': row['alternative'], 'score': float(row['score'])}
                                for _, row in df_result.iterrows()
                            ]
                        }
                    
                    # Sheet 4: Risk Assessment & Optimization (optional)
                    severity_values = {}
                    recommendations = {}
                    optimum_weights = {}
                    if 'RiskAssessment' in xls.sheet_names:
                        file_stream.seek(0)
                        df_risk = pd.read_excel(file_stream, sheet_name='RiskAssessment')
                        for _, row in df_risk.iterrows():
                            if 'alternative' in row and pd.notna(row['alternative']):
                                alt = str(row['alternative']).strip()
                                # Load severity if available
                                if 'severity' in row and pd.notna(row['severity']):
                                    try:
                                        severity_values[alt] = float(row['severity'])
                                    except (ValueError, TypeError):
                                        pass
                                # Load recommendation if available
                                if 'recommendation' in row and pd.notna(row['recommendation']):
                                    recommendations[alt] = str(row['recommendation']).strip()
                                # Load optimum weight if available
                                if 'optimum_weight' in row and pd.notna(row['optimum_weight']):
                                    try:
                                        optimum_weights[alt] = float(row['optimum_weight'])
                                    except (ValueError, TypeError):
                                        pass
                    
                    # Combine project and ranking data
                    if project_rows and len(project_rows) > 0:
                        # Has both project and ranking data
                        content = {
                            'rows': project_rows,
                            'columns': project_columns,
                            'criteriaWeights': criteria_weights,
                            'alternativesScores': alternatives_scores,
                            'rankingResult': ranking_result,
                            'rankingColumns': ranking_columns,
                            'groups': ranking_groups,
                            'severityValues': severity_values,
                            'recommendations': recommendations,
                            'optimumWeights': optimum_weights,
                            'criteriaDirections': criteria_directions
                        }
                    else:
                        # Only ranking data
                        content = {
                            'criteriaWeights': criteria_weights,
                            'alternativesScores': alternatives_scores,
                            'rankingResult': ranking_result,
                            'columns': ranking_columns,
                            'groups': ranking_groups,
                            'severityValues': severity_values,
                            'recommendations': recommendations,
                            'optimumWeights': optimum_weights,
                            'criteriaDirections': criteria_directions
                        }
                else:
                    # Regular Excel file - read first sheet (could be Project sheet)
                    file_stream.seek(0)
                    try:
                        # Try 'Project' sheet first
                        df = pd.read_excel(file_stream, sheet_name='Project')
                    except:
                        # Fallback to first sheet
                        file_stream.seek(0)
                        df = pd.read_excel(file_stream)
                    content = df.to_dict('records')
            except Exception as e:
                # Fallback to regular Excel reading
                file_stream.seek(0)
                df = pd.read_excel(file_stream)
                content = df.to_dict('records')
        elif file_extension == 'csv':
            file_stream.seek(0)
            df = pd.read_csv(file_stream)
            content = df.to_dict('records')
        elif file_extension == 'json':
            file_stream.seek(0)
            import json
            content = json.load(file_stream)
        else:
            return jsonify({'success': False, 'error': 'Unsupported file type'}), 400
        
        # Convert to rows/columns format if needed
        if isinstance(content, dict):
            # Check if it's library format (has 'headers')
            if 'headers' in content:
                formatted_content = content  # Library format
            # Check if it's ranking format (has 'criteriaWeights' or 'alternativesScores')
            elif 'criteriaWeights' in content or 'alternativesScores' in content:
                formatted_content = content  # Ranking format
            elif 'rows' in content:
                formatted_content = content  # Project format
            else:
                # Try to convert list to project format
                if isinstance(content, list) and len(content) > 0:
                    columns = list(content[0].keys())
                    rows = content
                    formatted_content = {
                        'rows': rows,
                        'columns': columns
                    }
                else:
                    formatted_content = {
                        'rows': [],
                        'columns': []
                    }
        elif isinstance(content, list) and len(content) > 0:
            columns = list(content[0].keys())
            rows = content
            formatted_content = {
                'rows': rows,
                'columns': columns
            }
        else:
            formatted_content = {
                'rows': content if isinstance(content, list) else [],
                'columns': list(content[0].keys()) if isinstance(content, list) and len(content) > 0 else []
            }
        
        return jsonify({
            'success': True,
            'data': formatted_content,
            'filename': filename
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@file_bp.route('/save', methods=['POST'])
def save():
    """Save project or ranking data as Excel file - returns file for download"""
    data = request.json
    file_data = data.get('data')
    filename = data.get('filename', 'project.xlsx')
    
    # Ensure filename has .xlsx extension
    if not filename.endswith('.xlsx'):
        filename = filename + '.xlsx'
    
    try:
        # Check if it has ranking data (has criteriaWeights or alternativesScores)
        has_ranking_data = isinstance(file_data, dict) and ('criteriaWeights' in file_data or 'alternativesScores' in file_data)
        # Check if it has project data (has rows)
        has_project_data = isinstance(file_data, dict) and 'rows' in file_data and file_data.get('rows')
        
        if has_ranking_data:
            # It has ranking data - extract ranking fields
            criteria_weights = file_data.get('criteriaWeights', {})
            alternatives_scores = file_data.get('alternativesScores', {})
            ranking_result = file_data.get('rankingResult', None)
            # Use rankingColumns if provided (when combined with project data), otherwise use columns (ranking-only data)
            ranking_columns = file_data.get('rankingColumns', []) if has_project_data else file_data.get('columns', [])
            ranking_groups = file_data.get('groups', []) if has_project_data else file_data.get('groups', [])
            # Extract Risk Assessment & Optimization data
            severity_values = file_data.get('severityValues', {})
            recommendations = file_data.get('recommendations', {})
            optimum_weights = file_data.get('optimumWeights', {})
            criteria_directions = file_data.get('criteriaDirections', {})
            
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Save project data first if available
                if has_project_data:
                    rows = file_data.get('rows', [])
                    project_columns = file_data.get('columns', [])
                    
                    # Project columns are already separate from ranking columns
                    project_cols_filtered = project_columns if isinstance(project_columns, list) else []
                    
                    df = pd.DataFrame(rows)
                    
                    # Determine column order: rowNo first, then project columns
                    if project_cols_filtered:
                        col_order = ['rowNo'] + [c for c in project_cols_filtered if c != 'rowNo']
                    elif len(rows) > 0:
                        all_keys = list(rows[0].keys())
                        if 'rowNo' in all_keys:
                            other_keys = [k for k in all_keys if k != 'rowNo']
                            col_order = ['rowNo'] + other_keys
                        else:
                            col_order = all_keys
                    else:
                        col_order = None
                    
                    # Reorder DataFrame columns if order is determined
                    if col_order:
                        # Ensure all columns from col_order exist in df
                        for col in col_order:
                            if col not in df.columns:
                                df[col] = pd.NA
                        # Get final column order: rowNo first, then ordered columns, then any remaining
                        final_cols = []
                        if 'rowNo' in df.columns:
                            final_cols.append('rowNo')
                        for col in col_order:
                            if col != 'rowNo' and col in df.columns:
                                final_cols.append(col)
                        # Add any remaining columns not in col_order
                        for col in df.columns:
                            if col not in final_cols:
                                final_cols.append(col)
                        df = df[final_cols]
                    
                    df.to_excel(writer, index=False, sheet_name='Project')
                
                # Save ranking sheets
                # Sheet: Ranking Columns
                if ranking_columns:
                    df_columns = pd.DataFrame([
                        {'column': col}
                        for col in ranking_columns
                    ])
                    df_columns.to_excel(writer, sheet_name='RankingColumns', index=False)
                
                # Sheet: Ranking Groups
                if ranking_groups:
                    groups_data = []
                    for group in ranking_groups:
                        if isinstance(group, dict) and 'name' in group and 'columns' in group and group['columns']:
                            groups_data.append({
                                'group_name': group['name'],
                                'columns': ','.join(group['columns'])
                            })
                    if groups_data:
                        df_groups = pd.DataFrame(groups_data)
                        df_groups.to_excel(writer, sheet_name='RankingGroups', index=False)
                    else:
                        pd.DataFrame(columns=['group_name', 'columns']).to_excel(writer, sheet_name='RankingGroups', index=False)
                else:
                    pd.DataFrame(columns=['group_name', 'columns']).to_excel(writer, sheet_name='RankingGroups', index=False)
                
                # Sheet: Criteria Weights (with directions)
                if criteria_weights:
                    df_weights_data = []
                    for k, v in criteria_weights.items():
                        row = {'criteria': k, 'weight': v}
                        # Add direction if available
                        if k in criteria_directions:
                            row['direction'] = criteria_directions[k]
                        else:
                            row['direction'] = pd.NA
                        df_weights_data.append(row)
                    df_weights = pd.DataFrame(df_weights_data)
                    df_weights.to_excel(writer, sheet_name='CriteriaWeights', index=False)
                elif criteria_directions:
                    # Even if no weights, save directions
                    df_weights_data = []
                    for k, v in criteria_directions.items():
                        df_weights_data.append({
                            'criteria': k,
                            'weight': pd.NA,
                            'direction': v
                        })
                    df_weights = pd.DataFrame(df_weights_data)
                    df_weights.to_excel(writer, sheet_name='CriteriaWeights', index=False)
                
                # Sheet: Alternatives Scores - New format: table with alternative column + criteria columns
                if alternatives_scores:
                    # Get all unique criteria from all alternatives
                    all_criteria = set()
                    for alt, scores in alternatives_scores.items():
                        if scores and isinstance(scores, dict):
                            all_criteria.update(scores.keys())
                    
                    # Also include criteria from ranking_columns if available
                    if ranking_columns:
                        all_criteria.update(ranking_columns)
                    
                    all_criteria = sorted(list(all_criteria))
                    
                    # Helper function to extract numeric value from alternative key for sorting
                    def extract_alt_number(alt_key):
                        """Extract numeric value from alternative key (e.g., 'Alternative 1' -> 1)"""
                        import re
                        if isinstance(alt_key, str):
                            match = re.search(r'\d+', alt_key)
                            if match:
                                return int(match.group())
                        return 0
                    
                    # Sort alternatives by numeric value (Alternative 1, 2, 3...)
                    sorted_alternatives = sorted(alternatives_scores.keys(), key=lambda x: extract_alt_number(str(x)))
                    
                    # Build table structure: alternative column + criteria columns
                    scores_rows = []
                    for alt in sorted_alternatives:
                        row = {'alternative': str(alt)}
                        scores = alternatives_scores.get(alt, {})
                        for criteria in all_criteria:
                            score_value = scores.get(criteria, '')
                            if score_value is not None and score_value != '':
                                try:
                                    row[criteria] = float(score_value)
                                except (ValueError, TypeError):
                                    row[criteria] = ''
                            else:
                                row[criteria] = ''
                        scores_rows.append(row)
                    
                    if scores_rows:
                        df_scores = pd.DataFrame(scores_rows)
                        df_scores.to_excel(writer, sheet_name='AlternativesScores', index=False)
                    else:
                        # Empty structure with alternative column
                        df_scores = pd.DataFrame(columns=['alternative'])
                        df_scores.to_excel(writer, sheet_name='AlternativesScores', index=False)
                
                # Sheet: Ranking Result
                if ranking_result and ranking_result.get('ranking'):
                    df_result = pd.DataFrame(ranking_result['ranking'])
                    df_result.to_excel(writer, sheet_name='RankingResult', index=False)
                
                # Sheet: Risk Assessment & Optimization
                if severity_values or recommendations or optimum_weights:
                    risk_rows = []
                    # Get all alternatives from all three sources
                    all_alternatives = set()
                    all_alternatives.update(severity_values.keys())
                    all_alternatives.update(recommendations.keys())
                    all_alternatives.update(optimum_weights.keys())
                    
                    # Sort alternatives by numeric value
                    def extract_alt_number(alt_key):
                        """Extract numeric value from alternative key"""
                        import re
                        if isinstance(alt_key, str):
                            match = re.search(r'\d+', alt_key)
                            if match:
                                return int(match.group())
                        return 0
                    
                    sorted_alternatives = sorted(all_alternatives, key=lambda x: extract_alt_number(str(x)))
                    
                    for alt in sorted_alternatives:
                        row = {'alternative': str(alt)}
                        # Add severity if available
                        if alt in severity_values:
                            try:
                                row['severity'] = float(severity_values[alt])
                            except (ValueError, TypeError):
                                row['severity'] = pd.NA
                        else:
                            row['severity'] = pd.NA
                        
                        # Add recommendation if available
                        if alt in recommendations:
                            row['recommendation'] = str(recommendations[alt]).strip() if recommendations[alt] else pd.NA
                        else:
                            row['recommendation'] = pd.NA
                        
                        # Add optimum weight if available
                        if alt in optimum_weights:
                            try:
                                row['optimum_weight'] = float(optimum_weights[alt])
                            except (ValueError, TypeError):
                                row['optimum_weight'] = pd.NA
                        else:
                            row['optimum_weight'] = pd.NA
                        
                        risk_rows.append(row)
                    
                    if risk_rows:
                        df_risk = pd.DataFrame(risk_rows)
                        df_risk.to_excel(writer, sheet_name='RiskAssessment', index=False)
                    else:
                        pd.DataFrame(columns=['alternative', 'severity', 'recommendation', 'optimum_weight']).to_excel(
                            writer, sheet_name='RiskAssessment', index=False)
                else:
                    # Empty structure
                    pd.DataFrame(columns=['alternative', 'severity', 'recommendation', 'optimum_weight']).to_excel(
                        writer, sheet_name='RiskAssessment', index=False)
                
                # Apply center alignment to all cells in all sheets
                workbook = writer.book
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    for row in sheet.iter_rows():
                        for cell in row:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            output.seek(0)
        else:
            # Regular project data - convert to rows format
            rows = file_data.get('rows', []) if isinstance(file_data, dict) else file_data
            columns = file_data.get('columns', []) if isinstance(file_data, dict) else []
            
            # If saving project.xlsx, check if we should preserve ranking data
            if filename == 'project.xlsx':
                from pathlib import Path
                DATA_DIR = Path(__file__).parent.parent.parent / 'data'
                filepath = DATA_DIR / 'project.xlsx'
                
                # Try to get ranking data from file_data first, then from existing file
                ranking_columns = []
                criteria_weights = {}
                alternatives_scores_df = None
                ranking_result = None
                severity_values = {}
                recommendations = {}
                optimum_weights = {}
                criteria_directions = {}
                
                # First, try to get from file_data if available
                if isinstance(file_data, dict):
                    ranking_columns = file_data.get('rankingColumns', [])
                    criteria_weights = file_data.get('criteriaWeights', {})
                    alternatives_scores = file_data.get('alternativesScores', {})
                    ranking_result = file_data.get('rankingResult', None)
                    severity_values = file_data.get('severityValues', {})
                    recommendations = file_data.get('recommendations', {})
                    optimum_weights = file_data.get('optimumWeights', {})
                    criteria_directions = file_data.get('criteriaDirections', {})
                    
                    # Convert alternatives_scores to DataFrame format if available
                    if alternatives_scores:
                        all_criteria = set()
                        for alt, scores in alternatives_scores.items():
                            if scores and isinstance(scores, dict):
                                all_criteria.update(scores.keys())
                        if ranking_columns:
                            all_criteria.update(ranking_columns)
                        all_criteria = sorted(list(all_criteria))
                        
                        def extract_alt_number(alt_key):
                            import re
                            if isinstance(alt_key, str):
                                match = re.search(r'\d+', alt_key)
                                if match:
                                    return int(match.group())
                            return 0
                        
                        sorted_alternatives = sorted(alternatives_scores.keys(), key=lambda x: extract_alt_number(str(x)))
                        scores_rows = []
                        for alt in sorted_alternatives:
                            row = {'alternative': str(alt)}
                            scores = alternatives_scores.get(alt, {})
                            for criteria in all_criteria:
                                score_value = scores.get(criteria, '')
                                if score_value is not None and score_value != '':
                                    try:
                                        row[criteria] = float(score_value)
                                    except (ValueError, TypeError):
                                        row[criteria] = ''
                                else:
                                    row[criteria] = ''
                            scores_rows.append(row)
                        if scores_rows:
                            alternatives_scores_df = pd.DataFrame(scores_rows)
                    
                    # Convert ranking_result to DataFrame if available
                    if ranking_result and ranking_result.get('ranking'):
                        ranking_result = pd.DataFrame(ranking_result['ranking'])
                
                # If not in file_data, try to load from existing file
                if not ranking_columns and not criteria_weights and filepath.exists():
                    try:
                        xls = pd.ExcelFile(filepath)
                        
                        if 'RankingColumns' in xls.sheet_names:
                            df_columns = pd.read_excel(filepath, sheet_name='RankingColumns')
                            ranking_columns = [
                                str(row['column']).strip() 
                                for _, row in df_columns.iterrows() 
                                if 'column' in row and pd.notna(row['column'])
                            ]
                        
                        if 'CriteriaWeights' in xls.sheet_names:
                            df_weights = pd.read_excel(filepath, sheet_name='CriteriaWeights')
                            for _, row in df_weights.iterrows():
                                if 'criteria' in row and 'weight' in row:
                                    try:
                                        criteria = str(row['criteria']).strip()
                                        weight = row['weight']
                                        if pd.notna(criteria) and pd.notna(weight):
                                            criteria_weights[criteria] = float(weight)
                                        # Load direction if available
                                        if 'direction' in row and pd.notna(row['direction']):
                                            criteria_directions[criteria] = str(row['direction']).strip()
                                    except (ValueError, TypeError):
                                        continue
                        
                        if 'AlternativesScores' in xls.sheet_names:
                            alternatives_scores_df = pd.read_excel(filepath, sheet_name='AlternativesScores')
                        
                        if 'RankingResult' in xls.sheet_names:
                            ranking_result = pd.read_excel(filepath, sheet_name='RankingResult')
                        
                        if 'RiskAssessment' in xls.sheet_names:
                            df_risk = pd.read_excel(filepath, sheet_name='RiskAssessment')
                            for _, row in df_risk.iterrows():
                                if 'alternative' in row and pd.notna(row['alternative']):
                                    alt = str(row['alternative']).strip()
                                    if 'severity' in row and pd.notna(row['severity']):
                                        try:
                                            severity_values[alt] = float(row['severity'])
                                        except (ValueError, TypeError):
                                            pass
                                    if 'recommendation' in row and pd.notna(row['recommendation']):
                                        recommendations[alt] = str(row['recommendation']).strip()
                                    if 'optimum_weight' in row and pd.notna(row['optimum_weight']):
                                        try:
                                            optimum_weights[alt] = float(row['optimum_weight'])
                                        except (ValueError, TypeError):
                                            pass
                    except Exception:
                        pass
                
                # Create Excel file in memory with all sheets
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    # Save project data with correct column order
                    df = pd.DataFrame(rows)
                    
                    # Determine column order: rowNo first, then columns from frontend, then any remaining
                    # Priority: 1) columns from frontend, 2) existing Excel order, 3) infer from rows
                    if columns:
                        # Use columns from frontend (preserve order from site)
                        # Frontend sends columns without rowNo, so add it at the beginning
                        col_order = ['rowNo'] + [c for c in columns if c != 'rowNo']
                    elif filepath.exists():
                        # Try to get order from existing Excel file
                        try:
                            df_existing = pd.read_excel(filepath, sheet_name='Project')
                            existing_cols = list(df_existing.columns)
                            if 'rowNo' in existing_cols:
                                other_cols = [c for c in existing_cols if c != 'rowNo']
                                col_order = ['rowNo'] + other_cols
                            else:
                                col_order = existing_cols
                        except:
                            # Fallback: infer from rows
                            if len(rows) > 0:
                                all_keys = list(rows[0].keys())
                                if 'rowNo' in all_keys:
                                    other_keys = [k for k in all_keys if k != 'rowNo']
                                    col_order = ['rowNo'] + other_keys
                                else:
                                    col_order = all_keys
                            else:
                                col_order = None
                    else:
                        # No existing file, infer from rows
                        if len(rows) > 0:
                            all_keys = list(rows[0].keys())
                            if 'rowNo' in all_keys:
                                other_keys = [k for k in all_keys if k != 'rowNo']
                                col_order = ['rowNo'] + other_keys
                            else:
                                col_order = all_keys
                        else:
                            col_order = None
                    
                    # Reorder DataFrame columns if order is determined
                    if col_order:
                        # Ensure all columns from col_order exist in df
                        for col in col_order:
                            if col not in df.columns:
                                df[col] = pd.NA
                        # Get final column order: rowNo first, then ordered columns, then any remaining
                        final_cols = []
                        if 'rowNo' in df.columns:
                            final_cols.append('rowNo')
                        for col in col_order:
                            if col != 'rowNo' and col in df.columns:
                                final_cols.append(col)
                        # Add any remaining columns not in col_order
                        for col in df.columns:
                            if col not in final_cols:
                                final_cols.append(col)
                        df = df[final_cols]
                    
                    df.to_excel(writer, index=False, sheet_name='Project')
                    
                    # Preserve ranking sheets
                    if ranking_columns:
                        df_columns = pd.DataFrame([{'column': col} for col in ranking_columns])
                        df_columns.to_excel(writer, sheet_name='RankingColumns', index=False)
                    
                    # Sheet: Criteria Weights (with directions)
                    if criteria_weights:
                        df_weights_data = []
                        for k, v in criteria_weights.items():
                            row = {'criteria': k, 'weight': v}
                            # Add direction if available
                            if k in criteria_directions:
                                row['direction'] = criteria_directions[k]
                            else:
                                row['direction'] = pd.NA
                            df_weights_data.append(row)
                        df_weights = pd.DataFrame(df_weights_data)
                        df_weights.to_excel(writer, sheet_name='CriteriaWeights', index=False)
                    elif criteria_directions:
                        # Even if no weights, save directions
                        df_weights_data = []
                        for k, v in criteria_directions.items():
                            df_weights_data.append({
                                'criteria': k,
                                'weight': pd.NA,
                                'direction': v
                            })
                        df_weights = pd.DataFrame(df_weights_data)
                        df_weights.to_excel(writer, sheet_name='CriteriaWeights', index=False)
                    
                    if alternatives_scores_df is not None and not alternatives_scores_df.empty:
                        alternatives_scores_df.to_excel(writer, sheet_name='AlternativesScores', index=False)
                    
                    if ranking_result is not None and not ranking_result.empty:
                        ranking_result.to_excel(writer, sheet_name='RankingResult', index=False)
                    
                    # Sheet: Risk Assessment & Optimization
                    if severity_values or recommendations or optimum_weights:
                        risk_rows = []
                        # Get all alternatives from all three sources
                        all_alternatives = set()
                        all_alternatives.update(severity_values.keys())
                        all_alternatives.update(recommendations.keys())
                        all_alternatives.update(optimum_weights.keys())
                        
                        # Sort alternatives by numeric value
                        def extract_alt_number(alt_key):
                            import re
                            if isinstance(alt_key, str):
                                match = re.search(r'\d+', alt_key)
                                if match:
                                    return int(match.group())
                            return 0
                        
                        sorted_alternatives = sorted(all_alternatives, key=lambda x: extract_alt_number(str(x)))
                        
                        for alt in sorted_alternatives:
                            row = {'alternative': str(alt)}
                            # Add severity if available
                            if alt in severity_values:
                                try:
                                    row['severity'] = float(severity_values[alt])
                                except (ValueError, TypeError):
                                    row['severity'] = pd.NA
                            else:
                                row['severity'] = pd.NA
                            
                            # Add recommendation if available
                            if alt in recommendations:
                                row['recommendation'] = str(recommendations[alt]).strip() if recommendations[alt] else pd.NA
                            else:
                                row['recommendation'] = pd.NA
                            
                            # Add optimum weight if available
                            if alt in optimum_weights:
                                try:
                                    row['optimum_weight'] = float(optimum_weights[alt])
                                except (ValueError, TypeError):
                                    row['optimum_weight'] = pd.NA
                            else:
                                row['optimum_weight'] = pd.NA
                            
                            risk_rows.append(row)
                        
                        if risk_rows:
                            df_risk = pd.DataFrame(risk_rows)
                            df_risk.to_excel(writer, sheet_name='RiskAssessment', index=False)
                        else:
                            pd.DataFrame(columns=['alternative', 'severity', 'recommendation', 'optimum_weight']).to_excel(
                                writer, sheet_name='RiskAssessment', index=False)
                    else:
                        # Empty structure
                        pd.DataFrame(columns=['alternative', 'severity', 'recommendation', 'optimum_weight']).to_excel(
                            writer, sheet_name='RiskAssessment', index=False)
                    
                    # Apply center alignment to all cells in all sheets
                    workbook = writer.book
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        for row in sheet.iter_rows():
                            for cell in row:
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                output.seek(0)
            else:
                # Create Excel file in memory
                output = io.BytesIO()
                df = pd.DataFrame(rows)
                
                # Determine column order: rowNo first, then columns from frontend, then any remaining
                if columns:
                    # Use columns from frontend (preserve order from site)
                    # Frontend sends columns without rowNo, so add it at the beginning
                    col_order = ['rowNo'] + [c for c in columns if c != 'rowNo']
                    # Ensure all columns from col_order exist in df
                    for col in col_order:
                        if col not in df.columns:
                            df[col] = pd.NA
                    # Get final column order: rowNo first, then ordered columns, then any remaining
                    final_cols = []
                    if 'rowNo' in df.columns:
                        final_cols.append('rowNo')
                    for col in col_order:
                        if col != 'rowNo' and col in df.columns:
                            final_cols.append(col)
                    # Add any remaining columns not in col_order
                    for col in df.columns:
                        if col not in final_cols:
                            final_cols.append(col)
                    df = df[final_cols]
                elif len(rows) > 0:
                    # Infer from rows
                    all_keys = list(rows[0].keys())
                    if 'rowNo' in all_keys:
                        other_keys = [k for k in all_keys if k != 'rowNo']
                        col_order = ['rowNo'] + other_keys
                        # Ensure all columns exist
                        for col in col_order:
                            if col not in df.columns:
                                df[col] = pd.NA
                        df = df[col_order]
                
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Project Data')
                    
                    # Apply center alignment to all cells
                    workbook = writer.book
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        for row in sheet.iter_rows():
                            for cell in row:
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@file_bp.route('/save-json', methods=['POST'])
def save_json_file():
    """Save data as JSON file - returns file for download"""
    data = request.json
    json_data = data.get('data')
    filename = data.get('filename', 'data.json')
    
    # Ensure filename has .json extension
    if not filename.endswith('.json'):
        filename = filename + '.json'
    
    try:
        import json
        # Create JSON file in memory
        output = io.BytesIO()
        json_str = json.dumps(json_data, indent=2, ensure_ascii=False)
        output.write(json_str.encode('utf-8'))
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/json',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@file_bp.route('/list', methods=['GET'])
def list_project_files():
    """List all project files"""
    try:
        files = list_files()
        return jsonify({
            'success': True,
            'files': files
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@file_bp.route('/download/<filename>', methods=['GET'])
def download_file(filename):
    """Download a file"""
    filepath = DATA_DIR / filename
    if not filepath.exists():
        return jsonify({'success': False, 'error': 'File not found'}), 404
    
    return send_file(str(filepath), as_attachment=True)

